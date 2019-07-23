#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
  $Id:  $

  This file is part of the Matisse pipeline GUI series
  Copyright (C) 2017- Observatoire de la Côte d'Azur

  Created on 2018-03-15
  @author: ame

  This software is a computer program whose purpose is to produce a smart log
  for the MATISSE instrument.

  This software is governed by the CeCILL  license under French law and
  abiding by the rules of distribution of free software.  You can  use, 
  modify and/ or redistribute the software under the terms of the CeCILL
  license as circulated by CEA, CNRS and INRIA at the following URL
  "http://www.cecill.info". 

  As a counterpart to the access to the source code and  rights to copy,
  modify and redistribute granted by the license, users are provided only
  with a limited warranty  and the software's author,  the holder of the
  economic rights,  and the successive licensors  have only  limited
  liability. 

  In this respect, the user's attention is drawn to the risks associated
  with loading,  using,  modifying and/or developing or reproducing the
  software by the user in light of its specific status of free software,
  that may mean  that it is complicated to manipulate,  and  that  also
  therefore means  that it is reserved for developers  and  experienced
  professionals having in-depth computer knowledge. Users are therefore
  encouraged to load and test the software's suitability as regards their
  requirements in conditions enabling the security of their systems and/or 
  data to be ensured and,  more generally, to use and operate it in the 
  same conditions as regards security. 

  The fact that you are presently reading this means that you have had
  knowledge of the CeCILL license and that you accept its terms.
"""
# Import necessary files
from libAutoPipeline import matisseType
import wx
import os, stat
from ObjectListView import ObjectListView, ColumnDefn
from astropy.io import fits
from fitsheaderviewer import fitsHeaderViewer
import distutils.spawn
import sys
import pickle
import socket
import mat_fileDialog
from mat_fileDialog import mat_FileDialog
from mat_fileDialog import identifyFile
import threading
import time
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill
from mat_time_flux_plot import mat_time_flux_plot
from mat_plotRmnrecOpd import mat_plotRmnrecOpd
from mat_acq_plot import mat_acq_plot
import shutil
from datetime import datetime
from mat_show_rawdata import mat_show_rawdata 
import subprocess

# Set useful paths
fvpath    = distutils.spawn.find_executable("fv")
iconspath = os.path.join(os.path.dirname(__file__),"icons")

#various dir0 for debug
if socket.gethostname()=="pc-amepro":
    dir0="D:\\Documents\\Travail\\MATISSE\\"
else:
    #on wmtpipeline
    dir0="/data/RawDataMatisse/"
print ("working in directory : {0}".format(dir0))

cleanLog = False

def findHeaderKeyword(h,key):
    try :
        res = h[key]
    except:
        print ("keyword {0} not found".format(key))
        res = ""
    return res

    
###############################################################################

class mat_logData():
    def __init__(self,tplstart,tplid,target,fluxL,fluxN,progid,nbFiles,nexp,comment,firstFileData,status):
        self.tplstart    = tplstart
        self.tplid       = tplid
        self.target      = target
        self.fluxL       = fluxL
        self.fluxN       = fluxN
        self.progid      = progid
        self.nbFiles     = nbFiles
        self.nexp        = nexp
        self.comment     = comment
        self.listOfFiles = [firstFileData]
        self.status      = status
        self.isok = True
        self.ok = ""
 #------------------------------------------------------------------------------         
    def getCSV(self,delimiter=";"):
        
       
        listOfFilesTxt      = ""
        for f in self.listOfFiles:
            
            listOfFilesTxt += f.filename
            listOfFilesTxt += "#"
     
        isImageAcq=" "
        isFringeSearch=" "   
        isPhotometry = " "
        dil = " "
        din = " "
        chop = "F"
        scitype = " "
        modl = "F"
        modn = "F"
        seeing = " "
        tau0 = " "
        wl0=" "
        if self.tplid=="MATISSE_img_acq":
            tpltype="ACQ"
            isImageAcq="F"
            isFringeSearch="F"        
            for f in self.listOfFiles:
                seeing=f.seeing
                tau0=f.tau0
                if f.modl == True:
                    modl = "T"
                if f.modn == True:
                    modn = "T"
                if f.chop == "T":
                    chop = "T"
                if f.dprtype == "SEARCH,STD":
                    isFringeSearch = "T"
                    if f.band == "L":
                        dil = f.disp
                    elif f.band == "N":
                        din = f.disp 
                if f.dprtype == "STD":
                    isImageAcq = "T"
        elif self.tplid == "MATISSE_hyb_obs":
            tpltype="OBS"
            
            isPhotometry = "F"
            for f in self.listOfFiles:
                seeing=f.seeing
                tau0=f.tau0
                if f.wl0 != "":
                    wl0=f.wl0
                if f.shutters == False:
                    isPhotometry = "T"
                if f.modl == True:
                    modl = "T"
                if f.modn == True:
                    modn = "T"
                scitype = f.scitype
                if f.chop == "T":
                    chop = "T"
                if f.band == "L":
                    dil = f.disp
                elif f.band == "N":
                    din = f.disp 
        else:
            tpltype="OTHER"
    
        print("CSV formatting {0}=>{1}".format(self.tplstart,tpltype))
            
        if tpltype=="OTHER":
            return ""
          

        res= "{1}{0}{2}{0}{3}{0}{4}{0}{5}{0}{6}{0}{7}{0}{8}{0}{9}{0}{10}{0} \
              {11}{0}{12}{0}{13}{0}{14}{0}{15}{0}{16}{0}{17}{0}{18}{0}{19}{0}{20}".format(
            delimiter, self.target,self.fluxL,self.fluxN,tpltype,self.tplstart,dil,wl0,din,modl,modn,
            isImageAcq,isFringeSearch,isPhotometry,
            chop,self.nbFiles,self.nexp,scitype,seeing,tau0,self.comment)
        print(res)
            

        return res
    
###############################################################################
        
class mat_fileData():
      def __init__(self,filename,h):
        print("Entering mat_filedata")
        self.filename = filename
        self.doCatg   = matisseType(h)
        self.date     = findHeaderKeyword(h,'DATE')
        self.expno    = "{0}/{1}".format(
            findHeaderKeyword(h,'HIERARCH ESO TPL EXPNO'),
            findHeaderKeyword(h,'HIERARCH ESO TPL NEXP'))
            
        self.dit  = findHeaderKeyword(h,"HIERARCH ESO DET SEQ1 DIT")
        self.ndit = findHeaderKeyword(h,"HIERARCH ESO DET NDIT")                    
        det = findHeaderKeyword(h,"HIERARCH ESO DET CHIP NAME")

        self.modl=findHeaderKeyword(h,"HIERARCH ESO INS OML1 ST")
        self.modn=findHeaderKeyword(h,"HIERARCH ESO INS OMN1 ST")
        self.scitype = findHeaderKeyword(h, "HIERARCH ESO DPR CATG")
        self.wl0=""

        
        if det == "HAWAII-2RG":
            self.band="L"
            self.disp = findHeaderKeyword(h,"HIERARCH ESO INS DIL NAME")
            #self.pisp = findHeaderKeyword(h,"HIERARCH ESO INS PIL NAME") 
            shutter1 = findHeaderKeyword(h,"HIERARCH ESO INS BSL1 ST")
            shutter2 = findHeaderKeyword(h,"HIERARCH ESO INS BSL2 ST")
            shutter3 = findHeaderKeyword(h,"HIERARCH ESO INS BSL3 ST")
            shutter4 = findHeaderKeyword(h,"HIERARCH ESO INS BSL4 ST")
            self.shutters = shutter1 and shutter2 and shutter3 and shutter4
            self.mod = self.modl
            self.wl0= findHeaderKeyword(h,"HIERARCH ESO SEQ DIL WL0")

            
        elif det=="AQUARIUS":
            self.band="N"
            self.disp = findHeaderKeyword(h,"HIERARCH ESO INS DIN NAME") 
            #self.pisp = findHeaderKeyword(h,"HIERARCH ESO INS PIN NAME") 
            shutter1 = findHeaderKeyword(h,"HIERARCH ESO INS BSN1 ST")
            shutter2 = findHeaderKeyword(h,"HIERARCH ESO INS BSN2 ST")
            shutter3 = findHeaderKeyword(h,"HIERARCH ESO INS BSN3 ST")
            shutter4 = findHeaderKeyword(h,"HIERARCH ESO INS BSN4 ST")
            self.shutters = shutter1 and shutter2 and shutter3 and shutter4
            self.mod = self.modn
        else:
            self.band=""
            self.disp=""
            #self.pisp=""
            self.shutters=""
            self.mod = ""
            
        self.dprtype=findHeaderKeyword(h,"HIERARCH ESO DPR TYPE")
        self.seeing = findHeaderKeyword(h,"HIERARCH ESO ISS AMBI FWHM START")
        self.tau0 = findHeaderKeyword(h,"HIERARCH ESO ISS AMBI TAU0 START")*1000

      
        self.chop = findHeaderKeyword(h,"HIERARCH ESO ISS CHOP ST")
        print self.shutters

###############################################################################

class mat_InteractivePipeline(wx.Dialog):

    def __init__(self, parent, date):
        super(mat_InteractivePipeline, self).__init__(parent, title="MATISSE Log of {0}".format(date), size=(1390, 900))
        
        self.date = os.path.basename(os.path.realpath(date).rstrip('\\').rstrip('/'))
        print("The current directory is "+self.date)
        self.logfilename = os.path.join(dir0,"mat_logger_"+self.date+'.pkl')
        self.csvfilename = os.path.join(dir0,"mat_logger_"+self.date+'.txt')
        self.excelfilename = os.path.join(dir0,"mat_logger_"+self.date+'.xlsx')


        print("The log file name is "+self.logfilename)
        self.tplList    = []       
        self.tplListObj = []
        self.fileList   = []
        self.InitUI()
        self.Centre()
        self.Show()
        self.path=""
        self.selection = []
                    
#------------------------------------------------------------------------------

    def InitUI(self):
        panel = wx.Panel(self)

        font  = wx.SystemSettings.GetFont(wx.SYS_SYSTEM_FONT)
            
        font.SetPointSize(7)

        
        
        hbox   = wx.BoxSizer(wx.HORIZONTAL)  
        hbox2  = wx.BoxSizer(wx.HORIZONTAL)    
        vbox   = wx.BoxSizer(wx.VERTICAL)   
        vbox2  = wx.BoxSizer(wx.VERTICAL)
        optionsSizer = wx.BoxSizer(wx.VERTICAL)
       
        self.tplListWidget = ObjectListView(panel,wx.ID_ANY, style=wx.LC_REPORT)   
        cols=[ ColumnDefn("Tpl. Start","left",130,"tplstart",minimumWidth=20),
               ColumnDefn("Tpl. Id",   "left",155,"tplid",   minimumWidth=20),
               ColumnDefn("Target",   "left",65, "target",  minimumWidth=20),
               ColumnDefn("Flux L",   "left",55, "fluxL",  minimumWidth=20),      
               ColumnDefn("Flux N",   "left",55, "fluxN", minimumWidth=20),                             
               #ColumnDefn("PROG. ID","left",70,"progid",minimumWidth=20),                        
               ColumnDefn("nFiles",    "left",35, "nbFiles", minimumWidth=20),                         
               ColumnDefn("nExp",     "left",35, "nexp",    minimumWidth=20),
               ColumnDefn(" ",      "left",30, "ok",   minimumWidth=20,checkStateGetter="isok")]
        self.tplListWidget.SetColumns(cols)
        self.tplListWidget.rowFormatter=self.setRowColorTpl
      
        
        self.tplListWidget.AutoSizeColumns()
        self.tplListWidget.SortBy(0, ascending=False)
        self.tplListWidget.SetFont(font)
              
        self.commentTxtCtrl= wx.TextCtrl(panel,style=wx.TE_MULTILINE)
        self.commentTxtCtrl.SetFont(font)
        self.fileListWidget =  ObjectListView(panel,wx.ID_ANY, style=wx.LC_REPORT)
        cols2=[ColumnDefn("Date",       "left",140,"date",      minimumWidth=20),
               ColumnDefn("File name",  "left",240,"filename",  minimumWidth=20),
               ColumnDefn("DO CATG",    "left",70, "doCatg",    minimumWidth=20),
               ColumnDefn("Disp.",      "left",40, "disp",      minimumWidth=20),
               ColumnDefn("wl0",        "left",40, "wl0",      minimumWidth=20),               
               ColumnDefn("DIT",        "left",40, "dit",       minimumWidth=20),
               ColumnDefn("NDIT",       "left",40, "ndit",      minimumWidth=20),
               ColumnDefn("Mod",        "left",35, "mod",       minimumWidth=20),
               ColumnDefn("Chp",        "left",30, "chop",      minimumWidth=20),
               ColumnDefn("expNo",      "left",37, "expno",     minimumWidth=20),
               ColumnDefn("Tau0",       "left",40, "tau0",      minimumWidth=20),
               ColumnDefn("Seeing",     "left",40, "seeing",    minimumWidth=20)]

        self.fileListWidget.SetColumns(cols2)
        self.fileListWidget.rowFormatter=self.setRowColorFile
        self.fileListWidget.AutoSizeColumns()
        self.fileListWidget.SortBy(0, ascending=False)
        self.fileListWidget.SetFont(font)
        
        
        self.updateBut    = wx.Button(panel,label="Update",style=wx.BU_EXACTFIT)
        self.exportCSVBut = wx.Button(panel,label="Export Excel",style=wx.BU_EXACTFIT)
        self.reduceSelectedBut    = wx.Button(panel,label="Reduce Selected",style=wx.BU_EXACTFIT)
        self.resetSelection    = wx.Button(panel,label="Reset Selection",style=wx.BU_EXACTFIT)        
        self.byebyeBut    = wx.Button(panel,label="Exit",style=wx.BU_EXACTFIT)
     
               
        vbox2.Add(self.fileListWidget, proportion=3, flag=wx.LEFT|wx.RIGHT|wx.EXPAND,border=5)
        vbox2.Add(optionsSizer, proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND,border=5)
        
        
        hbox.Add(self.tplListWidget,  proportion=67, flag=wx.LEFT|wx.RIGHT|wx.EXPAND,border=5)  
        hbox.Add(vbox2,               proportion=90,flag=wx.LEFT|wx.RIGHT|wx.EXPAND);
        









        
        hbox2.Add(self.updateBut,         proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)  
        hbox2.Add(self.exportCSVBut,      proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)  
        hbox2.Add(self.reduceSelectedBut, proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)  
        hbox2.Add(self.resetSelection,    proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)        
        hbox2.Add(self.byebyeBut,         proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)  

        
        
        vbox.Add(hbox,proportion=1,flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        vbox.Add(hbox2,              proportion=0.1,flag=wx.LEFT|wx.RIGHT|wx.EXPAND)

        
        
        panel.SetSizer(vbox)
         
        self.Bind(wx.EVT_LIST_ITEM_SELECTED,   self.tplSelected,   self.tplListWidget)
        self.Bind(wx.EVT_LIST_ITEM_SELECTED,   self.updateClicked, self.updateBut)
        self.Bind(wx.EVT_LIST_ITEM_SELECTED,   self.byebye,        self.byebyeBut)
        self.Bind(wx.EVT_TEXT,                 self.commentChanged,self.commentTxtCtrl)
        self.Bind(wx.EVT_LIST_ITEM_RIGHT_CLICK,self.fileListRightClicked,self.fileListWidget)
        self.Bind(wx.EVT_LIST_ITEM_RIGHT_CLICK,self.tplListRightClicked,self.tplListWidget)       
        self.exportCSVBut.Bind(wx.EVT_BUTTON,  self.exportCSVClicked)
        self.updateBut.Bind(wx.EVT_BUTTON,     self.updateClicked)
        self.reduceSelectedBut.Bind(wx.EVT_BUTTON,     self.reduceSelectedClicked)
        self.resetSelection.Bind(wx.EVT_BUTTON,     self.resetSelectionClicked)
        self.byebyeBut.Bind(wx.EVT_BUTTON,     self.byebye)
      
        if cleanLog==False:
            if os.path.isfile(self.logfilename):
                print "log file exist for {0} ... loading".format(self.logfilename)
                pik=open(self.logfilename, 'rb')
                self.tplList    = pickle.load(pik)
                self.tplListObj = pickle.load(pik)
                self.fileList   = pickle.load(pik)       
            else:
                print "No log file for {0} ...".format(self.date)
                         
        self.getInfosFromNight()

#------------------------------------------------------------------------------                
    def exportCSVClicked(self,event):
        print("Saving csv file to "+self.csvfilename)
        #csvfile = open(self.csvfilename,'wb')
        #csvfile.write("TARGET;TPLTYPE;TPLSTART;DIL;DIN;MODL;MODN;ACQ;FS;PHOTO;CHOPPING;NFILES;NEXP;TYPE;SEEING;TAU0;COMMENT\n");
      
        xl=Workbook()
        sheet=xl.active
        sheet.append(["TARGET","FLUXL","FLUXN","TPLTYPE","TPLSTART","DIL","WL0","DIN","MODL",
                    "MODN","ACQ","FS","PHOTO","CHOPPING","NFILES",
                      "NEXP","TYPE","SEEING","TAU0","COMMENT"])
        i=1
        newlist = sorted(self.tplListObj, key=lambda x: x.tplstart, reverse=False)
        for tplListObji in newlist :
            
            
            if tplListObji.getCSV():
                i+=1
                csvobj=tplListObji.getCSV()
                #csvfile.write(csvobj)
                xlobj=csvobj.split("\n")
                print(xlobj)
                nlines=len(xlobj)
                
                xlobj2=xlobj[0].split(";")
                
                
                sheet.append(xlobj2)
                if xlobj2[3]=="ACQ":
                     color = "8ce4ba"
                elif xlobj2[3]=="OBS":
                    color = "8db4e2"
                if tplListObji.isok == False:
                    color = "c86432"
                 
                print("{0} =>color={1}".format(xlobj2[3],color))
                
                for j in range(20):
                    cell=sheet['{0}{1}'.format(chr(65+j),i)]
                    cell.fill = PatternFill("solid",fgColor=color)
                
                for iline in range(1,nlines):
                    i+=1
                    xlobj2=[""]*19
                    xlobj2.append(xlobj[iline])
                    sheet.append(xlobj2)
                    for j in range(20):
                        cell=sheet['{0}{1}'.format(chr(65+j),i)]
                        cell.fill = PatternFill("solid",fgColor=color)
                    
            
        colwidth=[15,6,6,8,19,6,6,6,6,6,6,6,6,6,6,6,8,8,8,100]
        for i in range(20):
            sheet.column_dimensions["{0}".format(chr(65+i))].width = colwidth[i]
        #csvfile.close()

      

        xl.save(self.excelfilename)
        os.chmod(self.excelfilename, stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | 
                                   stat.S_IWGRP | stat.S_IROTH | stat.S_IWOTH )
 
         
#------------------------------------------------------------------------------                
    def tplListRightClicked(self,event):
        menu = wx.Menu()
        m1   = menu.Append( 0, "Add to selection" )
        menu.Bind(wx.EVT_MENU,self.addToSelection,m1) 
         
        self.tplListWidget.PopupMenu( menu, event.GetPoint())
 
          
#------------------------------------------------------------------------------                
    def fileListRightClicked(self,event):
        menu = wx.Menu()
        m1   = menu.Append( 0, "Show Header" )
        menu.Bind(wx.EVT_MENU,self.showHeader,m1)
        #wx.EVT_MENU( menu, 0, self.showHeader)
        m2   = menu.Append( 1, "Show RAW DATA")
        menu.Bind(wx.EVT_MENU,self.showRawData,m2)
        m3   = menu.Append( 2, "Plot Flux vs Time")
        menu.Bind(wx.EVT_MENU,self.plotFluxTime,m3)
        m4   = menu.Append( 3, "Open with fv")
        menu.Bind(wx.EVT_MENU,self.openWithFv,m4)
        m5   = menu.Append( 4, "Plot RMNREC OPD")
        menu.Bind(wx.EVT_MENU,self.plotRmnrecOpd,m5)
        m6   = menu.Append( 5, "Plot Acquisition")
        menu.Bind(wx.EVT_MENU,self.plotacq,m6)
        m7   = menu.Append( 6, "Copy files")
        menu.Bind(wx.EVT_MENU,self.copyFiles,m7) 
        m8   = menu.Append( 8, "Reduce data")
        menu.Bind(wx.EVT_MENU,self.reduceData,m8)
        m9   = menu.Append( 9, "Add to selection")
        menu.Bind(wx.EVT_MENU,self.addToSelection,m9)                
        #wx.EVT_MENU( menu, 1, self.showRawData)
        self.fileListWidget.PopupMenu( menu, event.GetPoint())
        
#---------cd /data/Tools/python---------------------------------------------------------------------


    def showHeader(self,event):

        itemNum  = self.fileListWidget.GetNextSelected(-1)            
        idx      = self.fileListWidget.GetItem(itemNum).GetData()
        l        = self.fileListWidget.GetObjects()
        filename = l[idx].filename
        print("Show header from  "+dir0+"/"+filename+"...")
        hv = fitsHeaderViewer(self,dir0+"/"+filename).Show()
                 
#------------------------------------------------------------------------------
                
    def showRawData(self,event):
        itemNum  = self.fileListWidget.GetNextSelected(-1)            
        idx      = self.fileListWidget.GetItem(itemNum).GetData()
        l        = self.fileListWidget.GetObjects()
        filename = l[idx].filename
        #dic      = mat_show_rawdata.open_mat(dir0+"/"+filename)
        print("Plotting data  from"+dir0+filename+"...")
        #mat_show_rawdata.show_mat(dic)
        mat_show_rawdata(dir0+"/"+filename)

#------------------------------------------------------------------------------
                
    def plotFluxTime(self,event):
        itemNum  = self.fileListWidget.GetNextSelected(-1)            
        idx      = self.fileListWidget.GetItem(itemNum).GetData()
        l        = self.fileListWidget.GetObjects()
        filename = l[idx].filename

        print("Plotting flux vs time for file "+ filename+"...")
        mat_time_flux_plot(filename)


#------------------------------------------------------------------------------
                
    def plotacq(self,event):
        itemNum  = self.fileListWidget.GetNextSelected(-1)            
        idx      = self.fileListWidget.GetItem(itemNum).GetData()
        l        = self.fileListWidget.GetObjects()
        filename = l[idx].filename

        print("Plotting Acquisition for file "+ filename+"...")
        mat_acq_plot(filename)


#------------------------------------------------------------------------------
                
    def openWithFv(self,event):
        itemNum  = self.fileListWidget.GetNextSelected(-1)            
        idx      = self.fileListWidget.GetItem(itemNum).GetData()
        l        = self.fileListWidget.GetObjects()
        filename = l[idx].filename
        print("Open {0} with fv".format(os.getcwd()+"/"+filename))
        subprocess.Popen(["fv" ,filename])


#------------------------------------------------------------------------------
                
    def plotRmnrecOpd(self,event):
        itemNum  = self.fileListWidget.GetNextSelected(-1)            
        idx      = self.fileListWidget.GetItem(itemNum).GetData()
        l        = self.fileListWidget.GetObjects()
        filename = l[idx].filename

        print("Plotting RMNREC OPD for file "+ filename+"...")
        mat_plotRmnrecOpd(filename,removeAvg=False,relative=False)


        
#------------------------------------------------------------------------------


    def copyFiles(self,event):
        l = self.fileListWidget.GetObjects()
        selectedFiles=[]
        itemNum  = self.fileListWidget.GetNextSelected(-1)
        while itemNum!=-1:
            idx = self.fileListWidget.GetItem(itemNum).GetData()
            selectedFiles.append(l[idx].filename)
            itemNum  = self.fileListWidget.GetNextSelected(itemNum)      

        dialog=wx.DirDialog(self,"Choose the directory")
        if dialog.ShowModal() == wx.ID_CANCEL:
            return

        path = dialog.GetPath()+"/"
        for filei in selectedFiles:
            shutil.copy(os.getcwd()+"/"+filei,path)


#------------------------------------------------------------------------------


    def reduceData(self,event):
        
        dialog=wx.DirDialog(self,"Choose an output directory")
        if dialog.ShowModal() == wx.ID_CANCEL:
            return
        path = dialog.GetPath()
        
        l = self.fileListWidget.GetObjects()
        selectedFiles=[]
        itemNum  = self.fileListWidget.GetNextSelected(-1)
        while itemNum!=-1:
            idx = self.fileListWidget.GetItem(itemNum).GetData()
            selectedFiles.append(l[idx].filename)
            itemNum  = self.fileListWidget.GetNextSelected(itemNum)      
        
        now=datetime.now().strftime("%H%M%S")
        os.mkdir(dir0+"/"+now)
        for filei in selectedFiles:
            shutil.copy(os.getcwd()+"/"+filei,dir0+"/"+now)
            

        

        command=("automaticPipeline.py --dirRaw={0} --dirResult={1} --nbCore=2;"
                "rm {0} -fr;"
                "cd {1};"
                "mat_tidyup_oifits.py {1}".format(os.getcwd()+"/"+now,path))
        
        print(command)

        subprocess.Popen(['xterm','-hold','-e',command])
        
        
        
        
      

#------------------------------------------------------------------------------

    def addToSelection(self,event):
          
        l = self.fileListWidget.GetObjects()
        selectedFiles=[]
        itemNum  = self.fileListWidget.GetNextSelected(-1)
        if itemNum!=-1:
            while itemNum!=-1:
                idx = self.fileListWidget.GetItem(itemNum).GetData()
                selectedFiles.append(os.getcwd()+"/"+l[idx].filename)
                itemNum  = self.fileListWidget.GetNextSelected(itemNum)      
        else:
           tpl = self.tplListWidget.GetObjects()  
           itemNum  = self.tplListWidget.GetNextSelected(-1)
           idx = self.tplListWidget.GetItem(itemNum).GetData()
           selectedFiles=[os.getcwd()+"/"+f.filename for f in tpl[idx].listOfFiles]
        print(selectedFiles)
        self.selection.extend(selectedFiles)
              
        
        
        
        
        
        
        
        
#------------------------------------------------------------------------------

    def saveData(self):
        pik = open(self.logfilename, 'wb')
        os.chmod(self.logfilename, stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | 
                                   stat.S_IWGRP | stat.S_IROTH | stat.S_IWOTH )

       
        pickle.dump(self.tplList, pik, pickle.HIGHEST_PROTOCOL)
        pickle.dump(self.tplListObj, pik, pickle.HIGHEST_PROTOCOL)
        pickle.dump(self.fileList, pik, pickle.HIGHEST_PROTOCOL)
        
#------------------------------------------------------------------------------
        
    def updateClicked(self,event):
        self.getInfosFromNight()
        self.tplListWidget.SortBy(0, ascending=False)
        
#------------------------------------------------------------------------------
        
    def reduceSelectedClicked(self,event):       
        print(self.selection)
        
        dialog=wx.DirDialog(self,"Choose the directory")
        if dialog.ShowModal() == wx.ID_CANCEL:
            return

        path = dialog.GetPath()+"/"
        
        command=("automaticPipeline.py --filesRaw=\"{0}\" --dirResult={1} --nbCore=2;"
                 "cd {1}; mat_tidyup_oifits.py {1}".format(self.selection,path))
        
        print(command)
        subprocess.Popen(['xterm','-hold','-e',command])
#------------------------------------------------------------------------------
        
    def resetSelectionClicked(self,event):       
        self.selection=[]
              
#------------------------------------------------------------------------------
        
    def byebye(self,event):
        print("Bye bye!")
        self.Destroy()
        return 0
    
#------------------------------------------------------------------------------
                
    def commentChanged(self,event):
        itemNum = self.tplListWidget.GetNextSelected(-1)
        if (itemNum!=-1):
            txt = self.tplListWidget.GetItemText(itemNum)
            idx = self.tplList.index(txt)   
            self.tplListObj[idx].comment = self.commentTxtCtrl.GetValue()
            self.saveData()
#------------------------------------------------------------------------------
           
    def tplSelected(self,event):
        #nfiles=self.tplListWidget.GetSelectedItemCount()
        itemNum=self.tplListWidget.GetNextSelected(-1)
        txt = self.tplListWidget.GetItemText(itemNum)
        idx = self.tplList.index(txt)
        self.fileListWidget.SetObjects(self.tplListObj[idx].listOfFiles)
        self.commentTxtCtrl.SetValue(self.tplListObj[idx].comment)
        
#------------------------------------------------------------------------------

    def getInfosFromNight(self):
        files = os.listdir(dir0)  
        for filei in files:
            if not(filei in self.fileList):
                if filei.endswith(".fits"):
                    try:
                        print("Reading header")
                        h        = fits.getheader(dir0+"/"+filei)
                        print("Getting tplStart")
                        tplstart = findHeaderKeyword(h,'HIERARCH ESO TPL START')
                        print "{0} ==> {1}".format(filei,tplstart)
                        if (tplstart in self.tplList):
                            print("tplStart already started, adding file to list")
                            i = self.tplList.index(tplstart)
                            print(i)
                            print(len(self.tplListObj))
                            print("incrementing file number...")
                            try:
                                self.tplListObj[i].nbFiles += 1
                                print("done")
                            except:
                                print("failed but...")
                                self.tplListObj[i].nbFiles = 1;
                                print("passed")
                            print("Checking exposure number")
                            if self.tplListObj[i].nexp  < findHeaderKeyword(h, 'HIERARCH ESO TPL NEXP'):
                                self.tplListObj[i].nexp = findHeaderKeyword(h, 'HIERARCH ESO TPL NEXP')
                            print("Adding file to list")
                            self.tplListObj[i].listOfFiles.append(mat_fileData(filei,h))
                            self.fileList.append(filei)     
                        elif tplstart!="":
                            print("Not tplStart, starting a new list")
                            self.tplList.append(tplstart)                           
                            target= findHeaderKeyword(h,'HIERARCH ESO OBS TARG NAME')       
                            fluxL=findHeaderKeyword(h,'HIERARCH ESO SEQ TARG FLUX L')
                            fluxN=findHeaderKeyword(h,'HIERARCH ESO SEQ TARG FLUX N')                                                                                          
                            progid= findHeaderKeyword(h,'HIERARCH ESO OBS PROG ID')
                            tplid=findHeaderKeyword(h,'HIERARCH ESO TPL ID')  
                            nexp=findHeaderKeyword(h,'HIERARCH ESO TPL NEXP')  
                            
                            self.tplListObj.append(mat_logData(tplstart,tplid,target,fluxL,fluxN,progid,
                                       1,nexp," ",mat_fileData(filei,h),"Started"))
                            self.fileList.append(filei)
                        else :  
                            print ("skipping file without tplstart {0}".format(filei))
                    except:
                        print ("skipping file {0} : not valid fits file".format(filei))
                else:
                    print ("skipping file {0} : not fits file".format(filei))
        self.tplListWidget.SetObjects(self.tplListObj)
        self.saveData()        
#------------------------------------------------------------------------------

# Colour of text for tpl list
    def setRowColorTpl(self,listItem, data):

        if data.tplid == "MATISSE_hyb_obs":
                txtcol=wx.Colour(116,196,147)
        elif data.tplid == "MATISSE_img_acq":
                txtcol=wx.Colour(116,147,196)
        else:
                txtcol=wx.Colour(155,155,155)      
        listItem.SetBackgroundColour(txtcol)

    # Colour of text for file list
    def setRowColorFile(self,listItem, data):
        try:
            bkgcol = mat_fileDialog.matisseColor[data.doCatg]
            sm     = sum(bkgcol)
            if sm > 2*255:
                txtcol=wx.BLACK
            else:
                txtcol=wx.WHITE
        except:
            txtcol = wx.BLACK
            bkgcol = mat_fileDialog.matisseColor["UNKNOWN"]    
        listItem.SetTextColour(txtcol)
        listItem.SetBackgroundColour(bkgcol)            

#------------------------------------------------------------------------------

class saveBackup(object):
    def __init__(self, interval=300):
        self.interval = interval

        thread = threading.Thread(target=self.run, args=())
        thread.daemon = True                            # Daemonize thread
        thread.start()  
    def run(self):
        """ Method that runs forever """
        while True:
            # Do something
            print('Save backup or something like that')
            logfilename = self.logfilename
            print(logfilename)

            time.sleep(self.interval)
            
#------------------------------------------------------------------------------

class autoUpdate(object):
    def __init__(self, interval=1):
        self.interval = interval

        thread = threading.Thread(target=self.run, args=())
        thread.daemon = True                            # Daemonize thread
        thread.start()  
    def run(self):
        """ Method that runs forever """
        while True:
            # Do something
            print('Updating stuff...')
            logfilename = mat_logger.logfilename
            print(logfilename)

            time.sleep(self.interval)
            
###############################################################################

if __name__ == '__main__':
    dir0=[];
    # Scan the command line arguments
    listArg = sys.argv
    for idx,elt in enumerate(listArg):
        if ('--help' in elt):
            print ("Usage : mat_logger.py date")
            print ("options")
            print ("--dir0  : base directory ")
            print ("--clean : remove old log file")
            sys.exit(0)
        if len(listArg) == 2:
            try:
                dir0=listArg[idx+1]
            except:
                print " no argument for dir0"
        if (elt=='--clean'):
            print "cleaning Log..."
            cleanLog=True
            
    # If no argument is given, then open the file dialog to select directory
    if not dir0:
        app2 = wx.App()
        print("No input directory given, running file selector...")
        openFileDialog = mat_FileDialog(None, 'Open a night directory',"lmk,")
        if openFileDialog.ShowModal() == wx.ID_OK:
            dir0 = openFileDialog.GetPaths()[0]
            print(dir0)
        openFileDialog.Destroy()
        app2.MainLoop()
        app2.Destroy()
    
    try :
        app = wx.App()
    except:
        pass
    
    # Save a backup pkl file, just in case something happens
    #saveBackup()
    #autoUpdate()
    
    interactivePipeline = mat_InteractivePipeline(None,dir0)
    interactivePipeline.Show()
    #if openLogger.ShowModal() == wx.ID_OK:
    #    print ("OK")
    #openLogger.Destroy()
    app.MainLoop()
    #app.Destroy()
