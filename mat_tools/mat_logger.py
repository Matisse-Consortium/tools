#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
  This file is part of the Matisse pipeline GUI series
  Copyright (C) 2017- Observatoire de la Côte d'Azur

  Created on 2018-03-15
  @author: ame

  This software is a computer program whose purpose is to produce a smart log
  for the MATISSE instrument.

  This software is governed by the CeCILL  license under French law and
  abiding by the rules of distribution of free software.

  You can use, modify and/ or redistribute the software under the terms
  of the CeCILL license as circulated by CEA, CNRS and INRIA at the
  following URL "http://www.cecill.info". You have a copy of the licence
  in the LICENCE.md file.

  The fact that you are presently reading this means that you have had
  knowledge of the CeCILL license and that you accept its terms.
"""

# Import necessary files
from libAutoPipeline import matisseType
import wx, wx.html
import os, stat
from ObjectListView import ObjectListView, ColumnDefn,Filter
from astropy.io import fits
import distutils.spawn
import sys
import pickle
import socket
import mat_fileDialog
from mat_fileDialog import mat_FileDialog
from mat_fileDialog import identifyFile
import threading
import time
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill
from mat_showFluxVsTime import mat_showFluxVsTime
from mat_showDLOffset import mat_showDLOffset
from mat_showAcq import mat_showAcq
import shutil
from datetime import datetime
from mat_showRawData import mat_showRawData
import matplotlib.pyplot as plt
from mat_showOiData import mat_showOiData
import subprocess
import argparse
from tqdm import tqdm

# Set useful paths
fvpath    = distutils.spawn.find_executable("fv")
iconspath = os.path.join(os.path.dirname(__file__),"icons")


cleanLog = False

def findHeaderKeyword(h,key):
    try :
        res = h[key]
    except:
        #print ("keyword {0} not found".format(key))
        res = ""
    return res


###############################################################################

class headerKeyword(object):
    def __init__(self,key,value,comment):
        self.keyword=key
        self.value=value
        self.comment=comment

###############################################################################

class mat_showFitsHeader(wx.Frame):
    def __init__(self,parent,headerOrFilename):


        

        if type(headerOrFilename)==type("") or  type(headerOrFilename)==type(u""):
            header=fits.getheader(headerOrFilename)
            filename=headerOrFilename
        else:
            header=headerOrFilename
            filename = "Unkwnown"
        wx.Frame.__init__(self, parent, title=filename,size=(800,700))
        self.tableView=ObjectListView(self,wx.ID_ANY, style=wx.LC_REPORT|wx.SUNKEN_BORDER)
        self.tableView.rowFormatter=self.setRowColor
        cols=[ColumnDefn("Keyword","left",250,"keyword",minimumWidth=100),\
              ColumnDefn("Value","left",250,"value",minimumWidth=100),\
              ColumnDefn("Comment","left",290,"comment",minimumWidth=100)]
        self.tableView.SetColumns(cols)

        self.searchField = wx.SearchCtrl(self, -1, "", style=wx.TE_PROCESS_ENTER | wx.NO_BORDER)
        self.searchField.Bind(wx.EVT_SEARCHCTRL_SEARCH_BTN, self.OnSearchKeyDown)
        self.searchField.Bind(wx.EVT_TEXT, self.OnSearchKeyDown)

        self.searchFieldText=wx.StaticText(self,-1,label="Filtering :")


        self.keywords=[]
        for keyi in header.keys():
            self.keywords.append(headerKeyword(keyi,header[keyi],header.comments[keyi]))


        self.tableView.SetObjects( self.keywords)
        
        
        vbox = wx.BoxSizer(wx.VERTICAL)
        hbox=  wx.BoxSizer(wx.HORIZONTAL)
        hbox.Add(self.searchFieldText,proportion=1)
        hbox.Add(self.searchField,proportion=1)
        
        vbox.AddSizer(hbox)
        vbox.Add(self.tableView,proportion=10)
        self.SetSizer(vbox)
        

    def OnSearchKeyDown(self,Event):
        #self.searchIdx+=1
        searchTxt=self.searchField.GetValue()
        filt=Filter.TextSearch(self.tableView,columns=(),text=searchTxt)
        self.tableView.SetFilter(filt)
        self.tableView.RepopulateList()

    def setRowColor(self,listItem, data):

        c=listItem.GetId() % 2
        listItem.SetBackgroundColour(wx.Colour(250+3*c,250+3*c,250+3*c))

###############################################################################

class mat_logData():
    def __init__(self,tplstart,tplid,target,fluxL,fluxN,ftracker,progid,nbFiles,nexp,comment,firstFileData,status):
        self.tplstart    = tplstart
        self.tplid       = tplid
        self.target      = target
        self.fluxL       = fluxL
        self.fluxN       = fluxN
        self.ftracker	 = ftracker
        self.progid      = progid
        self.nbFiles     = nbFiles
        self.nexp        = nexp
        self.comment     = comment
        self.listOfFiles = [firstFileData]
        self.status      = status
        self.isok = True
        self.ok = ""
 #------------------------------------------------------------------------------
    def getCSV(self,delimiter=";"):


        listOfFilesTxt      = ""
        for f in self.listOfFiles:

            listOfFilesTxt += f.filename
            listOfFilesTxt += "#"

        isImageAcq=" "
        isFringeSearch=" "
        isPhotometry = " "
        dil = " "
        din = " "
        chop = "F"
        scitype = " "
        modl = "F"
        modn = "F"
        seeing = " "
        tau0 = " "
        wl0=" "
        ditL="."
        ditN="."
        if (self.tplid=="MATISSE_img_acq") or (self.tplid=="MATISSE_img_acq_ft") :
            if (self.tplid=="MATISSE_img_acq"):
                tpltype="ACQ"
            else:
                tpltype="ACQ-FT"
            isImageAcq="F"
            isFringeSearch="F"
            for f in self.listOfFiles:
                seeing=f.seeing
                tau0=f.tau0
                if f.modl == True:
                    modl = "T"
                if f.modn == True:
                    modn = "T"
                if f.chop == "T":
                    chop = "T"
                if f.dprtype == "SEARCH,STD":
                    isFringeSearch = "T"
                    if f.band == "L":
                        dil = f.disp
                    elif f.band == "N":
                        din = f.disp
                if f.dprtype == "STD":
                    isImageAcq = "T"
        elif self.tplid == "MATISSE_hyb_obs" or self.tplid == "MATISSE_hse_obs" or  
             self.tplid == "MATISSE_hyb_obs_ft" or  self.tplid == "MATISSE_hyb_obs_ft_test" or
             self.tplid == "MATISSE_hyb_obs_vis" or  self.tplid == "MATISSE_hyb_obs_coh":
            if  self.tplid == "MATISSE_hyb_obs" :
                tpltype="OBS-HYB"
            elif self.tplid == "MATISSE_hyb_obs_ft" or  "MATISSE_hyb_obs_ft_test" :
                tpltype="OBS-HYB-FT"
            else:
                tpltype="OBS-HSE"

            isPhotometry = "F"
            for f in self.listOfFiles:
                seeing=f.seeing
                tau0=f.tau0
                if f.wl0 != "":
                    wl0=f.wl0
                if f.shutters == False:
                    isPhotometry = "T"
                if f.modl == True:
                    modl = "T"
                if f.modn == True:
                    modn = "T"
                scitype = f.scitype
                if f.chop == "T":
                    chop = "T"
                if f.band == "L":
                    dil = f.disp
                elif f.band == "N":
                    din = f.disp
                if f.band=="L":
                    ditL=int(float(f.dit)*1000)/1000.
                elif f.band=="N":
                    ditN=int(float(f.dit)*1000)/1000.

        else:
            tpltype="OTHER"

        #print("CSV formatting {0}=>{1}".format(self.tplstart,tpltype))

        if tpltype=="OTHER":
            return ""


        res= "{1}{0}{2}{0}{3}{0}{4}{0}{5}{0}{6}{0}{7}{0}{8}{0}{9}{0}{10}{0}"\
             "{11}{0}{12}{0}{13}{0}{14}{0}{15}{0}{16}{0}{17}{0}{18}{0}{19}{0}"\
             "{20}{0}{21}{0}{22}{0}{23}".format(delimiter, self.target,self.fluxL,
              self.fluxN,self.ftracker,tpltype,self.tplstart,dil,wl0,ditL,din,ditN,
            modl,modn,isImageAcq,isFringeSearch,isPhotometry,chop,self.nbFiles,
            self.nexp,scitype,seeing,tau0,self.comment)
        #print(res)


        return res

###############################################################################

class mat_fileData():
      def __init__(self,filename,h):
        #print("Entering mat_filedata")
        self.filename = filename
        self.doCatg   = matisseType(h)
        self.date     = findHeaderKeyword(h,'DATE')
        self.expno    = "{0}/{1}".format(
            findHeaderKeyword(h,'HIERARCH ESO TPL EXPNO'),
            findHeaderKeyword(h,'HIERARCH ESO TPL NEXP'))

        self.dit  = findHeaderKeyword(h,"HIERARCH ESO DET SEQ1 DIT")
        self.ndit = findHeaderKeyword(h,"HIERARCH ESO DET NDIT")
        det = findHeaderKeyword(h,"HIERARCH ESO DET CHIP NAME")

        self.modl=findHeaderKeyword(h,"HIERARCH ESO INS OML1 ST")
        self.modn=findHeaderKeyword(h,"HIERARCH ESO INS OMN1 ST")
        self.scitype = findHeaderKeyword(h, "HIERARCH ESO DPR CATG")
        self.wl0=""


        if det == "HAWAII-2RG":
            self.band="L"
            self.disp = findHeaderKeyword(h,"HIERARCH ESO INS DIL NAME")
            #self.pisp = findHeaderKeyword(h,"HIERARCH ESO INS PIL NAME")
            shutter1 = findHeaderKeyword(h,"HIERARCH ESO INS BSL1 ST")
            shutter2 = findHeaderKeyword(h,"HIERARCH ESO INS BSL2 ST")
            shutter3 = findHeaderKeyword(h,"HIERARCH ESO INS BSL3 ST")
            shutter4 = findHeaderKeyword(h,"HIERARCH ESO INS BSL4 ST")
            self.shutters = shutter1 and shutter2 and shutter3 and shutter4
            self.mod = self.modl
            self.wl0= findHeaderKeyword(h,"HIERARCH ESO SEQ DIL WL0")


        elif det=="AQUARIUS":
            self.band="N"
            self.disp = findHeaderKeyword(h,"HIERARCH ESO INS DIN NAME")
            #self.pisp = findHeaderKeyword(h,"HIERARCH ESO INS PIN NAME")
            shutter1 = findHeaderKeyword(h,"HIERARCH ESO INS BSN1 ST")
            shutter2 = findHeaderKeyword(h,"HIERARCH ESO INS BSN2 ST")
            shutter3 = findHeaderKeyword(h,"HIERARCH ESO INS BSN3 ST")
            shutter4 = findHeaderKeyword(h,"HIERARCH ESO INS BSN4 ST")
            self.shutters = shutter1 and shutter2 and shutter3 and shutter4
            self.mod = self.modn
        else:
            self.band=""
            self.disp=""
            #self.pisp=""
            self.shutters=""
            self.mod = ""

        self.dprtype=findHeaderKeyword(h,"HIERARCH ESO DPR TYPE")
        self.seeing = findHeaderKeyword(h,"HIERARCH ESO ISS AMBI FWHM START")
        self.tau0 = findHeaderKeyword(h,"HIERARCH ESO ISS AMBI TAU0 START")*1000


        self.chop = findHeaderKeyword(h,"HIERARCH ESO ISS CHOP ST")
        #print self.shutters

###############################################################################

class mat_showReductionLog(wx.Frame):
    def __init__(self, parent, filename,refreshTime=2):
        self.nlines = 0
        self.refreshTime=refreshTime
        self.filename=filename
        super(mat_showReductionLog, self).__init__(parent,title="reduction",size=(1000,800))
        self.text = wx.TextCtrl(self, style=wx.ALIGN_LEFT|wx.TE_MULTILINE|wx.TE_CHARWRAP|wx.TE_READONLY)
        self.updateTextFromFile()
        self.Show()

    def updateTextFromFile(self):
        f=open(self.filename,"r")
        txt=f.readlines()
        f.close()
        new_nlines=len(txt)-self.nlines
        for iline in range(new_nlines):
            self.text.AppendText(txt[self.nlines+iline])
        self.nlines=len(txt)
        wx.CallLater(self.refreshTime*1000,self.updateTextFromFile)


###############################################################################

class mat_bandParamOptions(wx.Frame):
    def __init__(self, parent, band):
        super(mat_bandParamOptions, self).__init__(parent,title="Help",size=(1000,800))

        html = wx.html.HtmlWindow(self)
        html.SetPage(
            "<h2>Param{0} options</h2>"
            "<p>Put here mat_raw_estimates recipes parameters for the {0} detector</p>"
            "<p><b>the \"--\" before the parameter name should be removed and the parameters should be separated by \"/\"<b></p>"
            "<h3>List of parameters:</h3>"
            "<table style='vertical-align:top;'>"
            "<tr><td><b>Name</b></td>"
            "<td><b>Default Value</b></td>"
            "<td><b>Description</b></td><tr>"
            "<tr><td>compensate</td>"
            "<td>pb,cb,rb,nl,if,bp,od</td>"
            "<td style='text-align:justify'>Defines which kind of compensation should be applied (none = no compensation at all, all = all compensations possible, dd ="
            "detector specific defaults, pb = subtract pixel bias, gb = subtract global bias, cb = subtract detector channel bias, rb ="
            "subtract row bias, ct = subtract crosstalk, nl = nonlinearity compensation, if = divide by instrument flat, df = divide by"
            "detector flat, bp = bad pixel interpolation, el = convert to electrons, od = remove optical distortion).</td></tr>"
            "<tr><td>gain</td>"
            "<td>0.0</td>"
            "<td>Default conversion gain in [e-/DU].</td></tr>"
            "<tr><td>reduce</td>"
            "<td>TRUE</td>"
            "<td>Flag if the reference sub-windows should be removed from the result.</td></tr>"
            "<tr><td>ioi</td>"
            "<td>0,0</td>"
            "<td>images of interest: <first>,<count>.</td></tr>"
            "<tr><td>tartyp</td>"
            "<td>0</td>"
            "<td>TARTYP estimation (0 = none, 1 = N*S+U+N*T+U, 2 = show intensity, 4 = show correlation, 8 = estimate TARTYP, 16 = change"
            "TIME, TARTYP, LOCALOPD and STEPPING_PHASE, 32 = exchange U with S or T)</td></tr>"
            "<tr><td>excess_count_lm</td>"
            "<td>0</td>"
            "<td>Excess frames mistakenly produced before first TIM-Board trigger(LM-Band).</td></tr>"
            "<tr><td>excess_count_n</td>"
            "<td>0</td>"
            "<td>Excess frames mistakenly produced before first TIM-Board trigger (N-Band).</td></tr>"
            "<tr><td>hampelFilterKernel</td>"
            "<td>0</td>"
            "<td>Only for L/M band. Apply atemporal Hampel filter to all pixels before deriving the photometry."
            "This filter improves the photometric estimation in case of faint stars (< 5Jy in L band with AT)."
            "The parameter fixes the size fo the kernel of the filter (kernel=10 is recommended).</td></tr>"
            "<tr><td>replaceTel</td>"
            "<td>0</td>"
            "<td>Replace Photometry of one telescope by the mean of the 3 others. (0: none, 1: AT1/UT1, 2: AT2/UT2, 3: AT3/UT3, 4: AT4/UT4).</td></tr>"
            "<tr><td>useOpdMod</td><td>FALSE</td><td>useOpdMod option.</td></tr>"
            "<tr><td>coherentIntegTime</td><td>0.0</td><td>Specify a coherent integration time (in s)</td></tr>"
            "<tr><td>corrFlux</td><td>FALSE</td><td>corrFlux option.</td></tr>"
            "<tr><td>cumulBlock</td><td>FALSE</td><td>cumul all blocks of an OB.</td></tr>"
            "<tr><td>coherentAlgo</td><td>2</td><td>Estimation Algorithm (1: AMBER like Method, 2: CRAL Cohrent Integration Method).</td></tr>"
            "<tr><td>spectralBinning</td><td>(5 for L 7 for N)</td><td>Spectral binning in pixels. Should be an odd number. If an even integer "
            "is given, the pipeline will select the next odd integer</td></tr>"
            "</table><h3>Examples</h3>"
            "<p> paramN=/useOpdMod=TRUE/corrFlux=TRUE</p>"
            "<p> paramL=/cumulBlock=TRUE</p>".format(band)
        )



###############################################################################

class mat_pipelineOptions(wx.Panel):

    def __init__(self, parent):
        super(mat_pipelineOptions, self).__init__(parent)

        tabPipelineSizer = wx.BoxSizer(wx.VERTICAL)
        hboxPipelineButtons = wx.BoxSizer(wx.HORIZONTAL)
        hboxPipeline1 = wx.BoxSizer(wx.HORIZONTAL)
        hboxPipeline2 = wx.BoxSizer(wx.HORIZONTAL)
        hboxPipeline3 = wx.BoxSizer(wx.HORIZONTAL)
        hboxPipeline4 = wx.BoxSizer(wx.HORIZONTAL)
        hboxPipeline5 = wx.BoxSizer(wx.HORIZONTAL)

        dirResultLabel = wx.StaticText(self, -1, "dirResult",style=wx.ALIGN_CENTER_HORIZONTAL)
        self.dirResultCtrl = wx.TextCtrl(self)
        dirResultBut = wx.Button(self,label="Choose",style=wx.BU_EXACTFIT)
        hboxPipeline1.Add(dirResultLabel,proportion=1, flag=wx.LEFT|wx.RIGHT| wx.ALIGN_CENTER_VERTICAL)
        hboxPipeline1.Add(self.dirResultCtrl,proportion=6, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        hboxPipeline1.Add(dirResultBut,proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)

        dirCalibLabel = wx.StaticText(self, -1, "dirCalib",style=wx.ALIGN_CENTER_HORIZONTAL)
        self.dirCalibCtrl = wx.TextCtrl(self)
        dirCalibBut = wx.Button(self,label="Choose",style=wx.BU_EXACTFIT)
        hboxPipeline2.Add(dirCalibLabel,proportion=1, flag=wx.LEFT|wx.RIGHT| wx.ALIGN_CENTER_VERTICAL)
        hboxPipeline2.Add(self.dirCalibCtrl,proportion=6, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        hboxPipeline2.Add(dirCalibBut,proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)

        nbCoreLabel = wx.StaticText(self, -1, "nbCore",style=wx.ALIGN_CENTER_HORIZONTAL)
        self.nbCoreCtrl = wx.Slider(self, value = 2, minValue = 1, maxValue = 8, style=wx.LEFT|wx.RIGHT| wx.ALIGN_CENTER_VERTICAL|wx.SL_LABELS)
        self.skipLCtrl = wx.CheckBox(self,label="skipL")
        self.skipNCtrl = wx.CheckBox(self,label="skipN")

        hboxPipeline3.Add(nbCoreLabel,proportion=1, flag=wx.LEFT|wx.RIGHT| wx.ALIGN_CENTER_VERTICAL)
        hboxPipeline3.Add(self.nbCoreCtrl,proportion=4, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        hboxPipeline3.Add(self.skipLCtrl,proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        hboxPipeline3.Add(self.skipNCtrl,proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)


        paramLLabel = wx.StaticText(self, -1, "paramL",style=wx.ALIGN_CENTER_HORIZONTAL)
        self.paramLCtrl = wx.TextCtrl(self)
        paramLBut = wx.Button(self,label="Help",style=wx.BU_EXACTFIT)
        hboxPipeline4.Add(paramLLabel,proportion=1, flag=wx.LEFT|wx.RIGHT| wx.ALIGN_CENTER_VERTICAL)
        hboxPipeline4.Add(self.paramLCtrl,proportion=6, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        hboxPipeline4.Add(paramLBut,proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)

        paramNLabel = wx.StaticText(self, -1, "paramN",style=wx.ALIGN_CENTER_HORIZONTAL)
        self.paramNCtrl = wx.TextCtrl(self)
        paramNBut = wx.Button(self,label="Help",style=wx.BU_EXACTFIT)
        hboxPipeline5.Add(paramNLabel,proportion=1, flag=wx.LEFT|wx.RIGHT| wx.ALIGN_CENTER_VERTICAL)
        hboxPipeline5.Add(self.paramNCtrl,proportion=6, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        hboxPipeline5.Add(paramNBut,proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)


        self.reduceSelectedBut    = wx.Button(self,label="Reduce Selected",style=wx.BU_EXACTFIT)
        self.resetSelection       = wx.Button(self,label="Reset Selection",style=wx.BU_EXACTFIT)
        hboxPipelineButtons.Add(self.reduceSelectedBut, proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        hboxPipelineButtons.Add(self.resetSelection,    proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)

        tabPipelineSizer.Add(hboxPipeline1,flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        tabPipelineSizer.Add(hboxPipeline2,flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        tabPipelineSizer.Add(hboxPipeline3,flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        tabPipelineSizer.Add(hboxPipeline4,flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        tabPipelineSizer.Add(hboxPipeline5,flag=wx.LEFT|wx.RIGHT|wx.EXPAND)

        tabPipelineSizer.Add(hboxPipelineButtons,flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        self.SetSizerAndFit(tabPipelineSizer)

        dirResultBut.Bind(wx.EVT_BUTTON,  self.dirResultChooseClicked)
        dirCalibBut.Bind(wx.EVT_BUTTON,  self.dirCalibChooseClicked)
        paramLBut.Bind(wx.EVT_BUTTON,  self.paramLButClicked)
        paramNBut.Bind(wx.EVT_BUTTON,  self.paramNButClicked)

    def dirResultChooseClicked(self,event):
        dialog=wx.DirDialog(self,"Choose the output directory")
        if dialog.ShowModal() == wx.ID_CANCEL:
            return
        self.dirResultCtrl.SetValue(dialog.GetPath()+"/")

    def dirCalibChooseClicked(self,event):
        dialog=wx.DirDialog(self,"Choose the calib directory")
        if dialog.ShowModal() == wx.ID_CANCEL:
            return
        self.dirCalibCtrl.SetValue(dialog.GetPath()+"/")

    def paramLButClicked(self,event):
        dlg = mat_bandParamOptions(self, "L")
        dlg.Show()

    def paramNButClicked(self,event):
        dlg = mat_bandParamOptions(self, "N")
        dlg.Show()


###############################################################################

class mat_logger(wx.Dialog):

    def __init__(self, parent, date, cleanLog=False,xlsOnly=False):

        self.date = os.path.basename(os.path.realpath(date).rstrip('\\').rstrip('/'))
        #print("The current directory is "+self.date)
        self.logfilename = os.path.join(dir0,"mat_logger_"+self.date+'.pkl')
        self.csvfilename = os.path.join(dir0,"mat_logger_"+self.date+'.txt')
        self.excelfilename = os.path.join(dir0,"mat_logger_"+self.date+'.xlsx')


        print("The log file name is "+self.logfilename)
        self.tplList    = []
        self.tplListObj = []
        self.fileList   = []
        
        if cleanLog==False:
            if os.path.isfile(self.logfilename):
                try:
                    print("log file exist for {0} ... loading".format(self.logfilename))
                    pik=open(self.logfilename, 'rb')
                    self.tplList    = pickle.load(pik)
                    self.tplListObj = pickle.load(pik)
                    self.fileList   = pickle.load(pik)
                except:
                    print("log file seems corrupted... archiving it anyway".format(self.logfilename))
                    os.rename(self.logfilename,self.logfilename+"old")
            else:
                print("No previous log file for {0} ...".format(self.date))
        else:
            print("Deleting previous log for {0} ...".format(self.date))

        self.getInfosFromNight()
        
        
        if xlsOnly==True:
            self.exportXLS()
            sys.exit(0)
            
        else:
            super(mat_logger, self).__init__(parent, title="MATISSE Log of {0}".format(date), style=wx.DEFAULT_FRAME_STYLE, size=(1450, 800))
            self.InitUI()
            self.tplListWidget.SetObjects(self.tplListObj)
            self.Centre()
            self.Show()
            self.path=""
            self.selectedTpl=[]

#------------------------------------------------------------------------------

    def InitUI(self):
        panel = wx.Panel(self)

        font  = wx.SystemSettings.GetFont(wx.SYS_SYSTEM_FONT)

        font.SetPointSize(7)

        hbox   = wx.BoxSizer(wx.HORIZONTAL)
        hbox2  = wx.BoxSizer(wx.HORIZONTAL)
        vbox   = wx.BoxSizer(wx.VERTICAL)
        vbox2  = wx.BoxSizer(wx.VERTICAL)

        tabs = wx.Notebook(panel)
        tabLog=wx.Panel(tabs)
        tabs.AddPage(tabLog,"Log")
        self.tabPipeline=mat_pipelineOptions(tabs)
        tabs.AddPage(self.tabPipeline,"Pipeline options")
        tabSelectedList=wx.Panel(tabs)
        tabs.AddPage(tabSelectedList,"Selected Obs.")


        self.tplListWidget = ObjectListView(panel,wx.ID_ANY, style=wx.LC_REPORT)
        cols=[ ColumnDefn("Tpl. Start","left",130,"tplstart",minimumWidth=20),
               ColumnDefn("Tpl. Id",   "left",155,"tplid",   minimumWidth=20),
               ColumnDefn("Target",   "left",65, "target",  minimumWidth=20),
               ColumnDefn("Flux L",   "left",55, "fluxL",  minimumWidth=20),
               ColumnDefn("Flux N",   "left",55, "fluxN", minimumWidth=20),
               ColumnDefn("FT",       "left",55, "ftracker", minimumWidth=20),
               #ColumnDefn("PROG. ID","left",70,"progid",minimumWidth=20),
               ColumnDefn("nFiles",    "left",35, "nbFiles", minimumWidth=20),
               ColumnDefn("nExp",     "left",35, "nexp",    minimumWidth=20),
               ColumnDefn(" ",      "left",30, "ok",   minimumWidth=20,checkStateGetter="isok")]
        self.tplListWidget.SetColumns(cols)
        self.tplListWidget.rowFormatter=self.setRowColorTpl
        self.tplListWidget.AutoSizeColumns()
        self.tplListWidget.SortBy(0, ascending=False)
        self.tplListWidget.SetFont(font)

        tabLogSizer = wx.BoxSizer()
        self.commentTxtCtrl= wx.TextCtrl(tabLog,style=wx.TE_MULTILINE)
        self.commentTxtCtrl.SetFont(font)
        tabLogSizer.Add(self.commentTxtCtrl,proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        tabLog.SetSizer(tabLogSizer)

        tabSelectedListSizer= wx.BoxSizer()
        self.selectedListWidget = ObjectListView(tabSelectedList,wx.ID_ANY, style=wx.LC_REPORT)
        cols=[ ColumnDefn("Tpl. Start","left",130,"tplstart",minimumWidth=20),
               ColumnDefn("Tpl. Id",   "left",155,"tplid",   minimumWidth=20),
               ColumnDefn("Target",   "left",65, "target",  minimumWidth=20),
               ColumnDefn("Flux L",   "left",55, "fluxL",  minimumWidth=20),
               ColumnDefn("Flux N",   "left",55, "fluxN", minimumWidth=20),
               ColumnDefn("FT",       "left",55, "ftracker", minimumWidth=20),
               ColumnDefn("nFiles",    "left",35, "nbFiles", minimumWidth=20),
               ColumnDefn("nExp",     "left",35, "nexp",    minimumWidth=20)]
        self.selectedListWidget.SetColumns(cols)
        self.selectedListWidget.AutoSizeColumns()
        self.selectedListWidget.SortBy(0, ascending=False)
        self.selectedListWidget.SetFont(font)
        tabSelectedListSizer.Add(self.selectedListWidget,proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        tabSelectedList.SetSizer(tabSelectedListSizer)

        self.fileListWidget =  ObjectListView(panel,wx.ID_ANY, style=wx.LC_REPORT)
        cols2=[ColumnDefn("Date",       "left",140,"date",      minimumWidth=20),
               ColumnDefn("File name",  "left",200,"filename",  minimumWidth=20),
               ColumnDefn("DO CATG",    "left",70, "doCatg",    minimumWidth=20),
               ColumnDefn("Disp.",      "left",40, "disp",      minimumWidth=20),
               ColumnDefn("wl0",        "left",40, "wl0",      minimumWidth=20),
               ColumnDefn("Band",        "left",40, "band",      minimumWidth=20),               
               ColumnDefn("DIT",        "left",40, "dit",       minimumWidth=20),
               ColumnDefn("NDIT",       "left",40, "ndit",      minimumWidth=20),
               ColumnDefn("Mod",        "left",35, "mod",       minimumWidth=20),
               ColumnDefn("Chp",        "left",30, "chop",      minimumWidth=20),
               ColumnDefn("expNo",      "left",37, "expno",     minimumWidth=20),
               ColumnDefn("Tau0",       "left",40, "tau0",      minimumWidth=20),
               ColumnDefn("Seeing",     "left",40, "seeing",    minimumWidth=20)]

        self.fileListWidget.SetColumns(cols2)
        self.fileListWidget.rowFormatter=self.setRowColorFile
        self.fileListWidget.AutoSizeColumns()
        self.fileListWidget.SortBy(0, ascending=False)
        self.fileListWidget.SetFont(font)

        self.updateBut    = wx.Button(panel,label="Update",style=wx.BU_EXACTFIT)
        self.exportCSVBut = wx.Button(panel,label="Export Excel",style=wx.BU_EXACTFIT)
        self.byebyeBut    = wx.Button(panel,label="Exit",style=wx.BU_EXACTFIT)


        vbox2.Add(self.fileListWidget, proportion=30, flag=wx.LEFT|wx.RIGHT|wx.EXPAND,border=5)
        #vbox2.Add(self.commentTxtCtrl, proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND,border=5)
        vbox2.Add(tabs, proportion=14, flag= wx.ALL|wx.EXPAND,border=5)

        hbox.Add(self.tplListWidget,  proportion=77, flag=wx.LEFT|wx.RIGHT|wx.EXPAND,border=5)
        hbox.Add(vbox2,               proportion=90,flag=wx.LEFT|wx.RIGHT|wx.EXPAND)

        hbox2.Add(self.updateBut,         proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        hbox2.Add(self.exportCSVBut,      proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        hbox2.Add(self.byebyeBut,         proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)



        vbox.Add(hbox,proportion=1,flag=wx.LEFT|wx.RIGHT|wx.EXPAND)
        vbox.Add(hbox2,              proportion=0.1,flag=wx.LEFT|wx.RIGHT|wx.EXPAND)



        panel.SetSizer(vbox)

        self.Bind(wx.EVT_LIST_ITEM_SELECTED,   self.tplSelected,   self.tplListWidget)
        self.Bind(wx.EVT_LIST_ITEM_SELECTED,   self.updateClicked, self.updateBut)
        self.Bind(wx.EVT_LIST_ITEM_SELECTED,   self.byebye,        self.byebyeBut)
        self.Bind(wx.EVT_TEXT,                 self.commentChanged,self.commentTxtCtrl)
        self.Bind(wx.EVT_LIST_ITEM_RIGHT_CLICK,self.fileListRightClicked,self.fileListWidget)
        self.Bind(wx.EVT_LIST_ITEM_RIGHT_CLICK,self.tplListRightClicked,self.tplListWidget)
        self.exportCSVBut.Bind(wx.EVT_BUTTON,  self.exportCSVClicked)
        self.updateBut.Bind(wx.EVT_BUTTON,     self.updateClicked)
        self.tabPipeline.reduceSelectedBut.Bind(wx.EVT_BUTTON,     self.reduceSelectedClicked)
        self.tabPipeline.resetSelection.Bind(wx.EVT_BUTTON,     self.resetSelectionClicked)
        self.byebyeBut.Bind(wx.EVT_BUTTON,     self.byebye)
        tabs.Bind(wx.EVT_NOTEBOOK_PAGE_CHANGED, self.tabChanged)
        self.Bind(wx.EVT_CLOSE, self.byebye)

        self.Show()

        # But solve a sizing problem only found on linux (but I don't know why)
        self.tabPipeline.Hide()
        self.tabPipeline.Show()
#------------------------------------------------------------------------------
    def tabChanged(self,event):
        pass
        #print("Tab changed")

#------------------------------------------------------------------------------
    def exportCSVClicked(self,event):
        self.exportXLS()
#------------------------------------------------------------------------------
    def exportXLS(self):
        print("Saving xls file to "+ self.excelfilename)
        
        xl=Workbook()
        sheet=xl.active
        sheet.append(["TARGET","FLUXL","FLUXN","FTRACK","TPLTYPE","TPLSTART","DIL","WL0","DITL","DIN","DITN","MODL",
                    "MODN","ACQ","FS","PHOT","CHOP","NFILES",
                      "NEXP","TYPE","SEEING","TAU0","COMMENT"])
        i=1
        newlist = sorted(self.tplListObj, key=lambda x: x.tplstart, reverse=False)
        for tplListObji in newlist :


            if tplListObji.getCSV():
                i+=1
                csvobj=tplListObji.getCSV()
                #csvfile.write(csvobj)
                xlobj=csvobj.split("\n")
                #print(xlobj)
                nlines=len(xlobj)

                xlobj2=xlobj[0].split(";")
                #print(xlobj2)
                sheet.append(xlobj2)
                if (xlobj2[4][0:3]=="ACQ") or  (xlobj2[4][0:3]=="ACQ-FT"):
                     color = "8ce4ba"
                elif xlobj2[4][0:3]=="OBS":
                    color = "8db4e2"

                if tplListObji.isok == False:
                    color = "c86432"

                #print("{0} =>color={1}".format(xlobj2[3],color))

                for j in range(23):
                    cell=sheet['{0}{1}'.format(chr(65+j),i)]
                    cell.fill = PatternFill("solid",fgColor=color)

                for iline in range(1,nlines):
                    i+=1
                    xlobj2=[""]*22
                    xlobj2.append(xlobj[iline])
                    sheet.append(xlobj2)
                    for j in range(23):
                        cell=sheet['{0}{1}'.format(chr(65+j),i)]
                        cell.fill = PatternFill("solid",fgColor=color)


        colwidth=[15,7,7,10,10,19,6,6,6,6,6,6,6,6,6,6,6,6,6,8,8,8,100]
        for i in range(23):
            sheet.column_dimensions["{0}".format(chr(65+i))].width = colwidth[i]
        #csvfile.close()



        xl.save(self.excelfilename)
        os.chmod(self.excelfilename, stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP |
                                   stat.S_IWGRP | stat.S_IROTH | stat.S_IWOTH )


#------------------------------------------------------------------------------
    def tplListRightClicked(self,event):
        menu = wx.Menu()
        m1   = menu.Append( 0, "Add to selection" )
        menu.Bind(wx.EVT_MENU,self.addToSelection,m1)

        self.tplListWidget.PopupMenu( menu, event.GetPoint())

#------------------------------------------------------------------------------
    def fileListRightClicked(self,event):
        menu = wx.Menu()
        idx = 0;
        m1   = menu.Append( idx, "Show Header" )
        menu.Bind(wx.EVT_MENU,self.showHeader,m1)
        idx+=1;
        m2   = menu.Append( idx, "Show RAW DATA")
        menu.Bind(wx.EVT_MENU,self.showRawData,m2)
        idx+=1;
        m2b   = menu.Append( idx, "Show OI DATA")
        menu.Bind(wx.EVT_MENU,self.showOiData,m2b)
        idx+=1;
        m3   = menu.Append(idx, "Plot Flux vs Time")
        menu.Bind(wx.EVT_MENU,self.plotFluxTime,m3)
        idx+=1;
        m4   = menu.Append( idx, "Open with fv")
        menu.Bind(wx.EVT_MENU,self.openWithFv,m4)
        idx+=1;
        m5   = menu.Append( idx, "Plot RMNREC OPD")
        menu.Bind(wx.EVT_MENU,self.plotRmnrecOpd,m5)
        idx+=1;
        m6   = menu.Append( idx, "Plot Acquisition")
        menu.Bind(wx.EVT_MENU,self.plotacq,m6)
        idx+=1;
        m7   = menu.Append( idx, "Copy files")
        menu.Bind(wx.EVT_MENU,self.copyFiles,m7)
        self.fileListWidget.PopupMenu( menu, event.GetPoint())


#---------cd /data/Tools/python---------------------------------------------------------------------
    def showHeader(self,event):

        itemNum  = self.fileListWidget.GetNextSelected(-1)
        idx      = self.fileListWidget.GetItem(itemNum).GetData()
        l        = self.fileListWidget.GetObjects()
        filename = l[idx].filename
        print("Show header from  "+dir0+"/"+filename+"...")
        hv = mat_showFitsHeader(self,dir0+"/"+filename).Show()

#------------------------------------------------------------------------------

    def showRawData(self,event):
        itemNum  = self.fileListWidget.GetNextSelected(-1)
        idx      = self.fileListWidget.GetItem(itemNum).GetData()
        l        = self.fileListWidget.GetObjects()
        filename = l[idx].filename
        #dic      = mat_show_rawdata.open_mat(dir0+"/"+filename)
        print("Plotting data  from"+dir0+filename+"...")
        #mat_show_rawdata.show_mat(dic)
        mat_showRawData(dir0+"/"+filename)

#------------------------------------------------------------------------------

    def showOiData(self,event):
        itemNum  = self.fileListWidget.GetNextSelected(-1)
        idx      = self.fileListWidget.GetItem(itemNum).GetData()
        l        = self.fileListWidget.GetObjects()
        filename = l[idx].filename
        #dic      = mat_show_rawdata.open_mat(dir0+"/"+filename)
        print("Plotting data  from"+dir0+filename+"...")
        #mat_show_rawdata.show_mat(dic)
        mat_showOiData(dir0+"/"+filename)
        plt.show()

#------------------------------------------------------------------------------

    def plotFluxTime(self,event):
        itemNum  = self.fileListWidget.GetNextSelected(-1)
        idx      = self.fileListWidget.GetItem(itemNum).GetData()
        l        = self.fileListWidget.GetObjects()
        filename = l[idx].filename

        print("Plotting flux vs time for file "+ filename+"...")
        mat_showFluxVsTime(filename)


#------------------------------------------------------------------------------

    def plotacq(self,event):
        itemNum  = self.fileListWidget.GetNextSelected(-1)
        idx      = self.fileListWidget.GetItem(itemNum).GetData()
        l        = self.fileListWidget.GetObjects()
        filename = l[idx].filename

        print("Plotting Acquisition for file "+ filename+"...")
        mat_showAcq(filename)


#------------------------------------------------------------------------------

    def openWithFv(self,event):
        itemNum  = self.fileListWidget.GetNextSelected(-1)
        idx      = self.fileListWidget.GetItem(itemNum).GetData()
        l        = self.fileListWidget.GetObjects()
        filename = l[idx].filename
        print("Open {0} with fv".format(os.getcwd()+"/"+filename))
        subprocess.Popen(["fv" ,filename])


#------------------------------------------------------------------------------

    def plotRmnrecOpd(self,event):
        itemNum  = self.fileListWidget.GetNextSelected(-1)
        idx      = self.fileListWidget.GetItem(itemNum).GetData()
        l        = self.fileListWidget.GetObjects()
        filename = l[idx].filename

        print("Plotting RMNREC OPD for file "+ filename+"...")
        mat_showDLOffset(filename,removeAvg=False,relative=False)

#------------------------------------------------------------------------------


    def copyFiles(self,event):
        l = self.fileListWidget.GetObjects()
        selectedFiles=[]
        itemNum  = self.fileListWidget.GetNextSelected(-1)
        while itemNum!=-1:
            idx = self.fileListWidget.GetItem(itemNum).GetData()
            selectedFiles.append(l[idx].filename)
            itemNum  = self.fileListWidget.GetNextSelected(itemNum)

        dialog=wx.DirDialog(self,"Choose the directory")
        if dialog.ShowModal() == wx.ID_CANCEL:
            return

        path = dialog.GetPath()+"/"
        for filei in selectedFiles:
            shutil.copy(os.getcwd()+"/"+filei,path)




#------------------------------------------------------------------------------

    def addToSelection(self,event):
        tpl = self.tplListWidget.GetObjects()
        itemNum  = self.tplListWidget.GetNextSelected(-1)
        while (itemNum!=-1):
            #print(itemNum)
            idx = self.tplListWidget.GetItem(itemNum).GetData()
            if not(tpl[idx] in self.selectedTpl):
                self.selectedTpl.append(tpl[idx])
            itemNum  = self.tplListWidget.GetNextSelected(itemNum)
        self.selectedListWidget.SetObjects(self.selectedTpl)
#------------------------------------------------------------------------------

    def saveData(self):
        try :
            pik = open(self.logfilename, 'wb')
            os.chmod(self.logfilename, stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP |
                                       stat.S_IWGRP | stat.S_IROTH | stat.S_IWOTH )


            pickle.dump(self.tplList, pik, pickle.HIGHEST_PROTOCOL)
            pickle.dump(self.tplListObj, pik, pickle.HIGHEST_PROTOCOL)
            pickle.dump(self.fileList, pik, pickle.HIGHEST_PROTOCOL)
        except:
            print("Unable to write {0}".format(self.logfilename))

#------------------------------------------------------------------------------

    def updateClicked(self,event):
        self.getInfosFromNight()
        self.tplListWidget.SetObjects(self.tplListObj)
        self.tplListWidget.SortBy(0, ascending=False)

#------------------------------------------------------------------------------

    def reduceSelectedClicked(self,event):
        addText=""

        dirResult = self.tabPipeline.dirResultCtrl.GetValue()
        if dirResult == "":
            dialog=wx.DirDialog(self,"Choose the output directory")
            if dialog.ShowModal() == wx.ID_CANCEL:
                return
            dirResult = dialog.GetPath()+"/"

        dirCalib = self.tabPipeline.dirCalibCtrl.GetValue()
        if dirCalib != "":
                addText+="--dirCalib={0} ".format(dirCalib)


        nbCore=self.tabPipeline.nbCoreCtrl.GetValue()

        skipL=self.tabPipeline.skipLCtrl.GetValue()
        if skipL==True:
            addText+="--skipL "

        skipN=self.tabPipeline.skipNCtrl.GetValue()
        if skipN==True:
            addText+="--skipN "

        paramL = self.tabPipeline.paramLCtrl.GetValue()
        if paramL != "":
                addText+="--paramL={0} ".format(paramL)

        paramN = self.tabPipeline.paramNCtrl.GetValue()
        if paramN != "":
                addText+="--paramN={0} ".format(paramN)


        selection=[]
        for tpli in self.selectedTpl:
            selection.extend([os.getcwd()+"/"+f.filename for f in tpli.listOfFiles])
        command=("mat_autoPipeline.py \"{0}\" --maxIter=4 --dirResult={1} --nbCore={2} {3};"
                 "cd {1};mat_tidyupOiFits.py .;mat_tidyupCalibMap.py .".format(selection,dirResult,nbCore,addText))

        time=datetime.now().strftime("%Y%m%d%H%M%S")
        filename="{0}/mat_reductionScript_{1}.csh".format(dirResult,time)
        logname="{0}/mat_reductionlog_{1}.txt".format(dirResult,time)
        f=open(filename,"w")
        f.write(command)
        f.close()


        subprocess.Popen(['nohup', 'csh', filename], stdout=open(logname, 'w'),stderr=open('/dev/null', 'w'),preexec_fn=os.setpgrp )
        mat_showReductionLog(self,logname)

        #cd subprocess.Popen(['xterm','-hold','-e',command])


#------------------------------------------------------------------------------

    def resetSelectionClicked(self,event):
        self.selectedTpl=[]
        self.selectedListWidget.SetObjects(self.selectedTpl)

#------------------------------------------------------------------------------

    def byebye(self,event):
        #print("Bye bye!")
        self.Destroy()
        return 0

#------------------------------------------------------------------------------

    def commentChanged(self,event):
        itemNum = self.tplListWidget.GetNextSelected(-1)
        if (itemNum!=-1):
            txt = self.tplListWidget.GetItemText(itemNum)
            idx = self.tplList.index(txt)
            self.tplListObj[idx].comment = self.commentTxtCtrl.GetValue()
            self.saveData()
#------------------------------------------------------------------------------

    def tplSelected(self,event):
        #nfiles=self.tplListWidget.GetSelectedItemCount()
        itemNum=self.tplListWidget.GetNextSelected(-1)
        txt = self.tplListWidget.GetItemText(itemNum)
        idx = self.tplList.index(txt)
        self.fileListWidget.SetObjects(self.tplListObj[idx].listOfFiles)
        self.commentTxtCtrl.SetValue(self.tplListObj[idx].comment)

#------------------------------------------------------------------------------

    def getInfosFromNight(self):
        files = os.listdir(dir0)
        print("Processing {0} files...".format(len(files)))
        for filei in tqdm(files):
            if not(filei in self.fileList):
                if filei.endswith(".fits"):
                    try:
                        #print("Reading header")
                        h = fits.getheader(dir0+"/"+filei)
                        #print("Getting tplStart")
                        tplstart = findHeaderKeyword(h,'HIERARCH ESO TPL START')
                        #print "{0} ==> {1}".format(filei,tplstart)
                        if (tplstart in self.tplList):
                            #print("tplStart already started, adding file to list")
                            i = self.tplList.index(tplstart)
                            #print(i)
                            #print(len(self.tplListObj))
                            #print("incrementing file number...")
                            try:
                                self.tplListObj[i].nbFiles += 1
                                #print("done")
                            except:
                                #print("failed but...")
                                self.tplListObj[i].nbFiles = 1;
                                #print("passed")
                            #print("Checking exposure number")
                            if self.tplListObj[i].nexp  < findHeaderKeyword(h, 'HIERARCH ESO TPL NEXP'):
                                self.tplListObj[i].nexp = findHeaderKeyword(h, 'HIERARCH ESO TPL NEXP')
                            #print("Adding file to list")
                            self.tplListObj[i].listOfFiles.append(mat_fileData(filei,h))
                            self.fileList.append(filei)
                        elif tplstart!="":
                            #print("Not tplStart, starting a new list")
                            self.tplList.append(tplstart)
                            target= findHeaderKeyword(h,'HIERARCH ESO OBS TARG NAME')
                            fluxL=findHeaderKeyword(h,'HIERARCH ESO SEQ TARG FLUX L')
                            fluxN=findHeaderKeyword(h,'HIERARCH ESO SEQ TARG FLUX N')
                            ftracker=findHeaderKeyword(h,'HIERARCH ESO DEL FT SENSOR')

                            progid= findHeaderKeyword(h,'HIERARCH ESO OBS PROG ID')
                            tplid=findHeaderKeyword(h,'HIERARCH ESO TPL ID')
                            nexp=findHeaderKeyword(h,'HIERARCH ESO TPL NEXP')

                            self.tplListObj.append(mat_logData(tplstart,tplid,target,fluxL,fluxN,ftracker,progid,
                                       1,nexp," ",mat_fileData(filei,h),"Started"))
                            self.fileList.append(filei)
                        else :
                            #print ("skipping file without tplstart {0}".format(filei))
                            pass
                    except:
                        pass
                        #print ("skipping file {0} : not valid fits file".format(filei))
                else:
                    pass
                    #print ("skipping file {0} : not fits file".format(filei))
       
        self.saveData()
#------------------------------------------------------------------------------

# Colour of text for tpl list
    def setRowColorTpl(self,listItem, data):

        if data.tplid == "MATISSE_hyb_obs" or data.tplid == "MATISSE_hse_obs" or  data.tplid =="MATISSE_hyb_obs_ft" or  data.tplid == "MATISSE_hyb_obs_ft_test":
                txtcol=wx.Colour(116,196,147)
        elif data.tplid == "MATISSE_img_acq" or data.tplid == "MATISSE_img_acq_ft" :
                txtcol=wx.Colour(116,147,196)
        else:
                txtcol=wx.Colour(155,155,155)
        listItem.SetBackgroundColour(txtcol)

    # Colour of text for file list
    def setRowColorFile(self,listItem, data):
        try:
            bkgcol = mat_fileDialog.matisseColor[data.doCatg]
            sm     = sum(bkgcol)
            if sm > 2*255:
                txtcol=wx.BLACK
            else:
                txtcol=wx.WHITE
        except:
            txtcol = wx.BLACK
            bkgcol = mat_fileDialog.matisseColor["UNKNOWN"]
        listItem.SetTextColour(txtcol)
        listItem.SetBackgroundColour(bkgcol)

#------------------------------------------------------------------------------

class saveBackup(object):
    def __init__(self, interval=300):
        self.interval = interval

        thread = threading.Thread(target=self.run, args=())
        thread.daemon = True                            # Daemonize thread
        thread.start()
    def run(self):
        """ Method that runs forever """
        while True:
            # Do something
            print('Save backup or something like that')
            logfilename = self.logfilename
            print(logfilename)

            time.sleep(self.interval)

#------------------------------------------------------------------------------

class autoUpdate(object):
    def __init__(self, interval=1):
        self.interval = interval

        thread = threading.Thread(target=self.run, args=())
        thread.daemon = True                            # Daemonize thread
        thread.start()
    def run(self):
        """ Method that runs forever """
        while True:
            # Do something
            print('Updating stuff...')
            logfilename = mat_logger.logfilename
            print(logfilename)

            time.sleep(self.interval)

###############################################################################

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='mat_logger.py is MATISSE interactive logger and data reduction tool')
    
    parser.add_argument('dirRaw', default="",  \
    help='The path to the directory containing your raw data.')

    parser.add_argument('--clean', default=0,action='store_true',  \
    help='remove old log file')


    parser.add_argument('--xlsOnly', default="",  action='store_true',\
    help='directly export log to XSL file (no interactive GUI (Not Working yet)')

    
    try:
        args = parser.parse_args()
    except:
        print("\n\033[93mRunning mat_logger.py --help to be kind with you:\033[0m\n")
        parser.print_help()
        print("\n     Example : python mat_logger.py /data/rawData/2018-05-19")
        sys.exit(0)
        
    try :
        app = wx.App()
    except:
        pass
    
    dir0=os.path.abspath(args.dirRaw)
    os.chdir(dir0)
    
    openLogger = mat_logger(None,dir0,args.clean,args.xlsOnly)
    openLogger.Show()
    app.MainLoop()

    
        
  
 

  
